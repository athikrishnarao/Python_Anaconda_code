import openpyxl
import mysql.connector
database_connector=mysql.connector.connect(user='root', password='Aegon@123', host='localhost')
Cursor_object=database_connector.cursor()
Cursor_object.execute("CREATE DATABASE Employee_information")
Cursor_object.execute('''USE Employee_information''')
Employee_information='''CREATE TABLE Employee(
ID int not null ,
Location VARCHAR(50)not null,
FirstName VARCHAR(50) not null,
LastName VARCHAR(50) not null,
Designation VARCHAR(50) not null) '''
Cursor_object.execute(Employee_information)
sql='''INSERT INTO Employee (ID,Location,FirstName,LastName,Designation ) values (%s,%s,%s,%s,%s)'''
wb=openpyxl.load_workbook("C:\TestLeaf_Python\Week4\EmployeeDatabaseinformation.xlsx")
Employee_information=[]
Sheetobject=wb["Sheet1"]
for row in range(2,Sheetobject.max_row+1):
    for column in range(1,Sheetobject.max_column+1):
                Employee_information.append(Sheetobject.cell(row,column).value)
c=int(input("Enter column size"))
val=list(zip(Employee_information[::c],Employee_information[1::c],Employee_information[2::c],Employee_information[3::c],Employee_information[4::c]))
Cursor_object.executemany(sql,val)
database_connector.commit()
print(Cursor_object.rowcount)
sql1='''Select * From Employee'''
Cursor_object.execute(sql1)
print(Cursor_object.fetchall())
database_connector.close()
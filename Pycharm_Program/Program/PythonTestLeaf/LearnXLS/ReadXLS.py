# Step 1 : Install openpyxl: pip install openpyxl

# Read the data from excel
import openpyxl

# Load the workbook
wp= openpyxl.load_workbook(r'Test.xlsx')

# return active sheet name of given workbook
print(wp.active)

# return all sheet name of the given workbook
print(wp.sheetnames)

# Creating object from given sheet name
x = wp['Sheet1']

# For row count
print("rows count", x.max_row)

# For Column count
print("Col count", x.max_column)


print(x.cell(1,1).value)

for i in range(1,5):
    print(x.cell(i,1).value)

for row in range(1,x.max_row+1):
    for col in range (1,x.max_column):
        print(x.cell(row,col).value)
        
#Way2 to read the xls in xls index
print(x['B2'].value)

for row in x['A1':'C3']:
    for col in row:
        print(col.value)
    print('x'*25)


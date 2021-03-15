"""import openpyxl
wb=openpyxl.load_workbook("Test.xlsx")
ExcelSheetList=wb.sheetnames
print(ExcelSheetList)
Number_of_sheets=len(ExcelSheetList)
print(Number_of_sheets)
for i in range(0,Number_of_sheets):
    Sheetobject=wb[ExcelSheetList[i]]
    print(Sheetobject)
    for row in range(1,Sheetobject.max_row+1):
        for column in range(1,Sheetobject.max_column+1):
            if (Sheetobject.cell(row,1).value=='Krish'):
               print(Sheetobject.cell(row,column).value)"""

import openpyxl
wb=openpyxl.load_workbook("ReadExcel.xlsx")
ExcelSheetList=wb.sheetnames
Number_of_sheets=len(ExcelSheetList)
print(Number_of_sheets)
for i in range(1,Number_of_sheets):
    Sheetobject=wb[ExcelSheetList[i]]
    print(Sheetobject)
    for row in range(1,Sheetobject.max_row+1):
        print(row)
        for column in range(1,Sheetobject.max_column+1):
            if (Sheetobject.cell(row,1).value=='15/03/2021'):
               print(Sheetobject.cell(row,column).value)

"""import openpyxl
wb=openpyxl.load_workbook("ReadExcel.xlsx")
ExcelSheetList=wb.sheetnames
Sheetobject=wb["OrderList"]
for row in range(1,Sheetobject.max_row+1):
    for column in range(1,Sheetobject.max_column+1):
        if (Sheetobject.cell(1,column).value=='OrderDate'):
               print(Sheetobject.cell(row,column).value)"""

"""import openpyxl
wb=openpyxl.load_workbook("ReadExcel.xlsx")
SheetList=wb.sheetnames
Number_of_sheets =len(SheetList)
for i in range(1,Number_of_sheets):
    Sheetobject=wb[SheetList[i]]
    for row in range(1,Sheetobject.max_row+1):
        for column in range(1,Sheetobject.max_column+1):
            if (Sheetobject.cell(row,1).value=='02/03/2021'):
               Sheetobject.cell(row,column).value='15/03/2021'
               wb.save("ReadExcel.xlsx")"""